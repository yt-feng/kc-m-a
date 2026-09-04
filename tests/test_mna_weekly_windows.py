from __future__ import annotations

import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook

from mna_weekly_tracker.config import OUTPUT_COLUMNS
from mna_weekly_tracker.url_validation import validate_workbook
from mna_weekly_tracker.weekly_windows import (
    expected_weekly_windows,
    latest_scheduled_cutoff,
    missing_weekly_windows,
    window_from_dates,
    workbook_completion,
)


def write_complete_workbook(path: Path) -> None:
    workbook = Workbook()
    cases = workbook.active
    cases.title = "周度并购案例"
    cases.append(OUTPUT_COLUMNS)
    row = {column: "已披露" for column in OUTPUT_COLUMNS}
    row.update(
        {
            "序号": 1,
            "并购方": "测试买方",
            "目标方": "测试标的",
            "URL": "https://example.com/disclosure/transaction",
        }
    )
    cases.append([row[column] for column in OUTPUT_COLUMNS])
    raw = workbook.create_sheet("原始候选")
    raw.append(["URL"])
    raw.append(["https://example.com/news/transaction"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


class WeeklyWindowTests(unittest.TestCase):
    def test_cutoff_is_fixed_to_friday_at_five(self) -> None:
        tz = ZoneInfo("Asia/Shanghai")

        after_cutoff = latest_scheduled_cutoff(datetime(2026, 9, 4, 6, 30, tzinfo=tz))
        before_cutoff = latest_scheduled_cutoff(datetime(2026, 9, 4, 4, 30, tzinfo=tz))

        self.assertEqual(after_cutoff, datetime(2026, 9, 4, 5, 0, tzinfo=tz))
        self.assertEqual(before_cutoff, datetime(2026, 8, 28, 5, 0, tzinfo=tz))

    def test_explicit_window_requires_exactly_seven_days(self) -> None:
        with self.assertRaisesRegex(ValueError, "exactly 7 days"):
            window_from_dates("2026-08-14", "2026-08-28")

    def test_explicit_window_requires_friday_boundaries(self) -> None:
        with self.assertRaisesRegex(ValueError, "start and end on Friday"):
            window_from_dates("2026-08-13", "2026-08-20")

    def test_expected_windows_are_contiguous_and_canonical(self) -> None:
        latest_end = window_from_dates("2026-08-28", "2026-09-04").end

        windows = expected_weekly_windows(latest_end, recovery_start_date="2026-08-07")

        self.assertEqual(
            [window.filename for window in windows],
            [
                "并购案例一览_20260807_20260814.xlsx",
                "并购案例一览_20260814_20260821.xlsx",
                "并购案例一览_20260821_20260828.xlsx",
                "并购案例一览_20260828_20260904.xlsx",
            ],
        )

    def test_missing_window_is_discovered_between_valid_neighbors(self) -> None:
        latest_end = window_from_dates("2026-08-28", "2026-09-04").end
        expected = expected_weekly_windows(latest_end, recovery_start_date="2026-07-31")
        with tempfile.TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            for window in expected:
                if window.filename != "并购案例一览_20260814_20260821.xlsx":
                    write_complete_workbook(window.output_path(output_dir))

            missing = missing_weekly_windows(
                output_dir,
                latest_end,
                recovery_start_date="2026-07-31",
            )

        self.assertEqual(
            [window.filename for window in missing],
            ["并购案例一览_20260814_20260821.xlsx"],
        )

    def test_corrupt_workbook_is_not_counted_as_complete(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "并购案例一览_20260814_20260821.xlsx"
            path.write_text("not an xlsx", encoding="utf-8")

            complete, detail = workbook_completion(path)

        self.assertFalse(complete)
        self.assertIn("unreadable", detail)

    def test_empty_workbook_fails_validation(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "empty.xlsx"
            workbook = Workbook()
            workbook.active.title = "周度并购案例"
            workbook.active.append(OUTPUT_COLUMNS)
            workbook.create_sheet("原始候选").append(["URL"])
            workbook.save(path)

            result = validate_workbook(path)

        self.assertIn("no_case_rows", {issue["type"] for issue in result["issues"]})


if __name__ == "__main__":
    unittest.main()
