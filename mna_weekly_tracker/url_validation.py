"""Validate weekly M&A workbook links."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .sources_fixed import is_usable_article_url, unwrap_news_url

BUSINESS_SHEETS = ("周度并购案例", "原始候选")


def _cell_url_values(cell) -> list[tuple[str, str]]:
    values: list[tuple[str, str]] = []
    if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
        values.append(("value", cell.value.strip()))
    if cell.hyperlink and cell.hyperlink.target:
        values.append(("hyperlink", str(cell.hyperlink.target).strip()))
    return values


def _headers(ws) -> dict[str, int]:
    return {str(cell.value or "").strip(): cell.column for cell in ws[1]}


def validate_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True)
    issues: list[dict[str, str]] = []
    url_count = 0
    case_rows = 0
    case_url_nonempty = 0

    for sheet_name in BUSINESS_SHEETS:
        if sheet_name not in workbook.sheetnames:
            issues.append({"sheet": sheet_name, "cell": "-", "type": "missing_sheet", "url": ""})
            continue
        ws = workbook[sheet_name]
        headers = _headers(ws)
        url_col = headers.get("URL")
        if sheet_name == "周度并购案例" and not url_col:
            issues.append({"sheet": sheet_name, "cell": "-", "type": "missing_url_column", "url": ""})

        for row_idx in range(2, ws.max_row + 1):
            if sheet_name == "周度并购案例":
                case_rows += 1
                raw_url = str(ws.cell(row=row_idx, column=url_col).value or "").strip() if url_col else ""
                clean_url = unwrap_news_url(raw_url)
                if clean_url and is_usable_article_url(clean_url):
                    case_url_nonempty += 1
                else:
                    issues.append({"sheet": sheet_name, "cell": f"{ws.cell(row=row_idx, column=url_col).coordinate if url_col else row_idx}", "type": "empty_or_unusable_case_url", "url": raw_url})

            for cell in ws[row_idx]:
                for value_type, url in _cell_url_values(cell):
                    url_count += 1
                    clean_url = unwrap_news_url(url)
                    if not is_usable_article_url(clean_url):
                        issues.append({"sheet": sheet_name, "cell": cell.coordinate, "type": f"unusable_{value_type}_url", "url": url})

    return {
        "path": str(path),
        "url_count": url_count,
        "case_rows": case_rows,
        "case_url_nonempty": case_url_nonempty,
        "issue_count": len(issues),
        "issues": issues[:100],
    }


def _latest(paths: list[Path]) -> list[Path]:
    return [sorted(paths)[-1]] if paths else []


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate weekly M&A Excel business links.")
    parser.add_argument("workbooks", nargs="+", type=Path)
    parser.add_argument("--latest", action="store_true", help="Only validate the latest matching workbook.")
    parser.add_argument("--write-json", type=Path)
    args = parser.parse_args()

    paths = [path for path in args.workbooks if path.exists()]
    if args.latest:
        paths = _latest(paths)
    results = [validate_workbook(path) for path in paths]
    payload = {"workbook_count": len(results), "results": results}
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    print(text)
    if args.write_json:
        args.write_json.parent.mkdir(parents=True, exist_ok=True)
        args.write_json.write_text(text, encoding="utf-8")
    if not results or any(result["issue_count"] for result in results):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
