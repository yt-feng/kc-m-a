"""Case selection for weekly and backfill M&A analysis reports."""

from __future__ import annotations

import json
import logging
import os
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from mna_weekly_tracker.sources_fixed import is_usable_article_url, unwrap_news_url
from mna_weekly_tracker.sources_rich import RawItem, fetch_all_candidates

from .case_pool import EXTENDED_CASE_POOL
from .config import CASE_DISCOVERY_QUERIES, CATEGORIES, CLASSIC_CASE_SEEDS, TOPIC_SELECTION_RULES
from .deepseek_client import chat_json

LOGGER = logging.getLogger(__name__)
BEIJING_TZ = ZoneInfo("Asia/Shanghai")
VAGUE_PARTY_TERMS = (
    "未披露", "未知", "不详", "待定", "某标的", "标的资产", "标的公司", "相关资产", "部分资产",
    "旗下资产", "金融资产", "相关股权", "受让方", "转让方", "未具名", "控股子公司",
    "实施情况报告书", "相关事项", "交易对方", "交易各方", "本次交易", "本次收购",
)
PARTY_SUFFIX_RE = re.compile(
    r"(股份有限公司|有限责任公司|有限公司|控股集团|控股有限公司|控股|集团|公司|"
    r"corporation|inc\.?|limited|ltd\.?|plc|holdings?)",
    re.I,
)
PLACEHOLDER_TEXT = {"", "-", "无", "未知", "不详", "未披露", "n/a", "na", "none", "null"}
OFFICIAL_SOURCE_DOMAINS = (
    "cninfo.com.cn",
    "static.cninfo.com.cn",
    "sse.com.cn",
    "szse.cn",
    "bse.cn",
    "neeq.com.cn",
    "hkexnews.hk",
    "sec.gov",
    "samr.gov.cn",
    "csrc.gov.cn",
    "ndrc.gov.cn",
    "mofcom.gov.cn",
)
DEAL_NUMBER_RE = re.compile(
    r"\d+(?:,\d{3})*(?:\.\d+)?\s*(?:%|％|亿元|亿美元|亿港元|万港元|万元|美元|港元|元|股|股份|股权|"
    r"crore|million|billion|bn|mn)?",
    re.I,
)
DEAL_VALUE_RE = re.compile(
    r"\d+(?:,\d{3})*(?:\.\d+)?\s*(?:%|％|亿元|亿美元|亿港元|万港元|万元|美元|港元|元|股|股份|股权|"
    r"crore|million|billion|bn|mn)",
    re.I,
)
TRANSACTION_STRUCTURE_RE = re.compile(
    r"(全部股权|全部股份|控股权|控制权|表决权|支付现金|现金支付|现金收购|"
    r"发行股份(?:及支付现金)?|股份支付|定向发行|定增|全额认购|认购|"
    r"协议转让|要约收购|收购报告书|权益变动报告书|股权转让|股份转让|受让.{0,20}(?:股权|股份)|"
    r"非货币(?:出资|实缴)|作价出资)",
    re.I,
)
RICH_DISCLOSURE_HINTS = (
    "要约收购报告书", "要约收购", "权益变动报告书", "详式权益变动", "收购报告书",
    "重大资产重组报告书", "重组报告书", "交易报告书", "草案", "预案",
    "发行股份及支付现金", "协议转让", "控制权变更", "完成过户", "结果公告",
)
DETAIL_RICH_DISCLOSURE_HINTS = (
    "要约收购报告书", "权益变动报告书", "详式权益变动", "收购报告书",
    "重大资产重组报告书", "重组报告书", "交易报告书", "草案", "预案",
    "发行股份及支付现金", "协议转让",
)
THIN_DISCLOSURE_HINTS = ("完成过户", "结果公告", "期限届满", "停牌公告", "提示性公告")
COMPLETION_ANNOUNCEMENT_HINTS = (
    "完成过户", "过户完成", "完成交割", "交割完成", "交易完成", "收购完成", "实施完成",
    "要约收购结果", "结果公告", "股权交割", "工商变更登记手续",
)
CONTROL_COMPLETION_HINTS = (
    "控股股东发生变更", "控股股东变更", "控制权发生变更", "实际控制人发生变更", "成为公司控股股东",
    "变更实际控制人", "变更控股股东", "控股股东、实际控制人发生变更", "实际控制人、控股股东",
    "将成为上市公司控股股东", "将成为公司控股股东",
)
NON_CONTROL_HINTS = (
    "不会导致公司控股股东", "不会导致公司实际控制人", "不涉及要约收购", "实际控制权未发生变化",
)
COMPLETION_SKIP_TITLE_HINTS = (
    "独立财务顾问", "法律意见书", "核查意见", "减值测试", "财务顾问核查",
)
MATERIAL_SHARE_TRANSFER_HINTS = (
    "协议转让", "股份转让", "完成过户", "过户完成", "权益变动", "证券过户登记确认书",
)
ASSET_SWAP_HINTS = ("资产置换", "置入资产", "置出资产", "完成交割", "资产交割确认书")
TAVILY_COMPLETION_QUERIES: tuple[str, ...] = (
    "site:static.cninfo.com.cn 完成过户 控制权 股份转让 2026",
    "site:static.cninfo.com.cn 标的资产 过户完成 重大资产购买 2026",
    "site:static.cninfo.com.cn 完成交割 资产置换 股权 2026",
    "site:hkexnews.hk acquisition completion discloseable transaction 2026",
    "site:hkexnews.hk connected transaction acquisition completion 2026",
    "site:sec.gov acquisition completed merger closed 2026 8-K",
    "site:announcements.asx.com.au scheme implementation acquisition completed 2026",
    "site:businesswire.com acquisition completed closed deal value 2026",
)


@dataclass
class CaseBrief:
    case_name: str
    category: str
    region: str
    source_title: str = ""
    source_url: str = ""
    published_at: str = ""
    why: str = ""
    is_domestic: bool = False
    is_classic: bool = False
    completed_year: str = ""
    is_completed: bool = False
    acquirer: str = ""
    target: str = ""
    deal_value: str = ""
    deal_status: str = ""
    buyer_motivation: str = ""
    seller_motivation: str = ""
    financial_highlights: str = ""

    def key(self) -> str:
        return f"{self.case_name}|{self.category}".lower().strip()

    def identity_key(self) -> str:
        return case_identity_key(self)

    def is_recent_completed(self) -> bool:
        return self.is_completed and self.completed_year in {"2025", "2026"} and not has_unresolved_incomplete_signal(self.deal_status)

    def is_allowed_topic(self) -> bool:
        if self.is_recent_completed() or self.is_classic:
            return True
        return False

    def to_dict(self) -> dict[str, object]:
        return {
            "case_name": self.case_name,
            "category": self.category,
            "region": self.region,
            "source_title": self.source_title,
            "source_url": self.source_url,
            "published_at": self.published_at,
            "why": self.why,
            "is_domestic": self.is_domestic,
            "is_classic": self.is_classic,
            "completed_year": self.completed_year,
            "is_completed": self.is_completed,
            "acquirer": self.acquirer,
            "target": self.target,
            "deal_value": self.deal_value,
            "deal_status": self.deal_status,
            "buyer_motivation": self.buyer_motivation,
            "seller_motivation": self.seller_motivation,
            "financial_highlights": self.financial_highlights,
        }


def excluded_terms() -> list[str]:
    raw = os.getenv("REPORT_EXCLUDE_CASE_TERMS", "")
    return [term.strip().lower() for term in re.split(r"[,，;；\n]+", raw) if term.strip()]


def env_int(name: str, default: int, *, minimum: int | None = None) -> int:
    try:
        value = int(os.getenv(name, str(default)))
    except ValueError:
        value = default
    if minimum is not None:
        value = max(value, minimum)
    return value


def env_flag(name: str, *, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def source_notice(message: str) -> None:
    safe = str(message).replace("\n", " ")[:1000]
    if os.getenv("GITHUB_ACTIONS") == "true":
        print(f"::notice::{safe}", flush=True)


def clean_cell(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def is_placeholder(value: str | None) -> bool:
    return clean_cell(value).lower() in PLACEHOLDER_TEXT


def parse_bool(value: object, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    text = clean_cell(value).lower()
    if text in {"true", "1", "yes", "y", "是", "已完成", "完成"}:
        return True
    if text in {"false", "0", "no", "n", "否", "未完成", "进行中", "审批中", "意向", "终止"}:
        return False
    return default


def has_deal_number(value: str | None) -> bool:
    text = clean_cell(value)
    if is_placeholder(text):
        return False
    return bool(DEAL_NUMBER_RE.search(text))


def has_deal_value_signal(value: str | None) -> bool:
    text = clean_cell(value)
    if is_placeholder(text):
        return False
    return bool(DEAL_VALUE_RE.search(text) or TRANSACTION_STRUCTURE_RE.search(text))


def extract_transaction_terms(*values: str) -> str:
    text = clean_cell("；".join(value for value in values if value and not is_placeholder(value)))
    if not text:
        return ""
    terms: list[str] = []

    def add(label: str) -> None:
        if label not in terms:
            terms.append(label)

    for match in DEAL_VALUE_RE.finditer(text):
        add(match.group(0))
        if len(terms) >= 6:
            break
    if re.search(r"全部股权|全部股份", text):
        add("全部股权/股份")
    if re.search(r"支付现金|现金支付|现金收购", text):
        add("现金支付")
    if re.search(r"发行股份及支付现金", text):
        add("发行股份及支付现金")
    elif re.search(r"发行股份|股份支付", text):
        add("股份支付")
    if re.search(r"协议转让", text):
        add("协议转让")
    if re.search(r"股权转让|股份转让|受让.{0,20}(?:股权|股份)", text):
        add("股权/股份受让")
    if re.search(r"要约收购", text):
        add("要约收购")
    if re.search(r"定向发行|定增|全额认购|认购", text):
        add("定向发行/认购")
    if re.search(r"非货币(?:出资|实缴)|作价出资", text):
        add("非货币出资/作价出资")
    if re.search(r"控股权|控制权|表决权", text):
        add("控制权/表决权安排")
    return "；".join(terms[:8])


COMPLETED_STATUS_PATTERNS = (
    r"已(?:完成|交割|过户)",
    r"(?:交易|收购|合并|股权转让|股份转让|资产转让|资产过户|交割|过户)(?:已经|已)?完成",
    r"完成(?:交易|收购|并购|交割|过户|合并|股权转让|股份转让|资产转让|资产过户)",
    r"\bclosed\b",
    r"\bcompleted\b",
)
INCOMPLETE_STATUS_PATTERNS = (
    r"进行中",
    r"审批中",
    r"问询",
    r"上会",
    r"尚未",
    r"未完成",
    r"待(?:完成|交割|过户|审批|审核)",
    r"预计",
    r"拟",
    r"计划",
    r"将(?:于|在)?",
    r"\bpending\b",
    r"\bexpected\b",
    r"subject to",
    r"will close",
    r"to close",
    r"closing conditions?",
    r"regulatory approval",
)


def has_completed_signal(value: str | None) -> bool:
    text = clean_cell(value).lower()
    if not text:
        return False
    has_strong_completed = any(re.search(pattern, text, re.I) for pattern in COMPLETED_STATUS_PATTERNS)
    if not has_strong_completed:
        return False
    already_completed = any(
        re.search(pattern, text, re.I)
        for pattern in (
            r"已(?:完成|交割|过户)",
            r"(?:交易|收购|合并|股权转让|股份转让|资产转让|资产过户|交割|过户)(?:已经|已)?完成",
            r"完成(?:交易|收购|并购|交割|过户|合并|股权转让|股份转让|资产转让|资产过户)",
            r"\bclosed\b",
        )
    )
    if not already_completed and any(re.search(pattern, text, re.I) for pattern in INCOMPLETE_STATUS_PATTERNS):
        return False
    return True


def has_incomplete_signal(value: str | None) -> bool:
    text = clean_cell(value).lower()
    if not text:
        return False
    return any(re.search(pattern, text, re.I) for pattern in INCOMPLETE_STATUS_PATTERNS)


def has_unresolved_incomplete_signal(value: str | None) -> bool:
    return has_incomplete_signal(value) and not has_completed_signal(value)


def is_report_completed_candidate(brief: CaseBrief) -> bool:
    """True only for cases eligible for formal Word reports."""
    if has_unresolved_incomplete_signal(brief.deal_status):
        return False
    return brief.is_classic or brief.is_completed or has_completed_signal(brief.deal_status)


def has_usable_source_url(url: str | None) -> bool:
    cleaned = unwrap_news_url(clean_cell(url))
    return not is_placeholder(cleaned) and is_usable_article_url(cleaned)


def is_authoritative_source_url(url: str | None) -> bool:
    cleaned = unwrap_news_url(clean_cell(url))
    if not has_usable_source_url(cleaned):
        return False
    parsed = re.match(r"^https?://([^/?#]+)([^?#]*)", cleaned, re.I)
    if not parsed:
        return False
    host = parsed.group(1).lower().replace("www.", "")
    path = (parsed.group(2) or "").lower()
    return path.endswith(".pdf") or any(domain in host for domain in OFFICIAL_SOURCE_DOMAINS)


def is_report_ready_candidate(brief: CaseBrief) -> bool:
    """Cheap preflight before spending minutes on research and model calls."""
    if not has_explicit_parties(brief):
        return False
    if not is_report_completed_candidate(brief):
        return False
    if not has_usable_source_url(brief.source_url) and not brief.is_classic and not brief.is_recent_completed():
        return False
    evidence_text = "\n".join([
        brief.deal_value,
        brief.why,
        brief.source_title,
        brief.financial_highlights,
    ])
    if not has_deal_value_signal(evidence_text):
        if not (
            is_authoritative_source_url(brief.source_url)
            and (has_detail_rich_disclosure_signal(brief) or (brief.is_completed and has_rich_disclosure_signal(brief)))
        ):
            return False
    return True


def is_report_source_ready_candidate(brief: CaseBrief) -> bool:
    return is_report_ready_candidate(brief) and has_usable_source_url(brief.source_url)


def is_report_source_linked_completed_candidate(brief: CaseBrief) -> bool:
    return (
        has_explicit_parties(brief)
        and is_report_completed_candidate(brief)
        and has_usable_source_url(brief.source_url)
        and brief.is_allowed_topic()
        and not is_excluded_case(brief)
    )


def has_rich_disclosure_signal(brief: CaseBrief) -> bool:
    text = "\n".join([brief.source_title, brief.why, brief.deal_status, brief.source_url])
    return any(token in text for token in RICH_DISCLOSURE_HINTS)


def has_detail_rich_disclosure_signal(brief: CaseBrief) -> bool:
    text = "\n".join([brief.source_title, brief.why, brief.deal_status, brief.source_url])
    return any(token in text for token in DETAIL_RICH_DISCLOSURE_HINTS)


def report_candidate_priority(brief: CaseBrief) -> tuple[int, int, int, int, int, int, str]:
    disclosure_text = "\n".join([brief.source_title, brief.why, brief.deal_status])
    if any(token in disclosure_text for token in DETAIL_RICH_DISCLOSURE_HINTS):
        disclosure_penalty = 0
    elif any(token in disclosure_text for token in THIN_DISCLOSURE_HINTS):
        disclosure_penalty = 2
    else:
        disclosure_penalty = 1
    completed_penalty = 0 if is_report_completed_candidate(brief) else 10
    url_penalty = 0 if has_usable_source_url(brief.source_url) else 8
    deal_signal = has_deal_value_signal("\n".join([brief.deal_value, brief.financial_highlights, brief.why, brief.source_title]))
    if deal_signal:
        deal_penalty = 0
    elif is_authoritative_source_url(brief.source_url) and has_detail_rich_disclosure_signal(brief):
        deal_penalty = 1
    else:
        deal_penalty = 2 if is_authoritative_source_url(brief.source_url) else 5
    positive_control_signal = "控制权" in brief.case_name or (
        any(hint in disclosure_text for hint in CONTROL_COMPLETION_HINTS)
        and not any(hint in disclosure_text for hint in NON_CONTROL_HINTS)
    )
    if (
        re.search(r"\d+(?:\.\d+)?%股份", brief.case_name)
        and not positive_control_signal
        and not re.search(r"(?:购买资产|资产置换|发行股份)", disclosure_text)
    ):
        percent_values = _percent_values(brief.case_name)
        largest_percent = max(percent_values) if percent_values else 0
        deal_penalty += 4
        if largest_percent and largest_percent < 10:
            deal_penalty += 8
    rationale_penalty = 0
    if len(clean_cell(brief.buyer_motivation or brief.why)) < 15:
        rationale_penalty += 1
    if len(clean_cell(brief.seller_motivation)) < 15:
        rationale_penalty += 1
    classic_penalty = 2 if brief.is_classic else 0
    return (completed_penalty, deal_penalty, disclosure_penalty, url_penalty, rationale_penalty, classic_penalty, brief.case_name)


def raw_report_item_score(item: RawItem) -> tuple[int, str, str]:
    """Prioritize raw public-source items that are likely completed deal evidence."""
    url = unwrap_news_url(item.url)
    title_text = " ".join([item.title, item.summary, item.query])
    score = 0
    if is_authoritative_source_url(url):
        score += 1000
    if any(hint in title_text for hint in COMPLETION_ANNOUNCEMENT_HINTS):
        score += 500
    if any(hint in title_text for hint in CONTROL_COMPLETION_HINTS):
        score += 220
    if any(hint in title_text for hint in ASSET_SWAP_HINTS):
        score += 180
    if any(hint in title_text for hint in MATERIAL_SHARE_TRANSFER_HINTS):
        score += 120
    if has_completed_signal(title_text):
        score += 120
    if has_deal_value_signal(title_text):
        score += 40
    if any(hint in item.title for hint in COMPLETION_SKIP_TITLE_HINTS):
        score -= 300
    return score, item.published_at or "", item.title


def report_rejection_reason(brief: CaseBrief, historical_keys: set[str] | None = None) -> str:
    if historical_keys is not None and any_key_in_history(case_identity_keys(brief), historical_keys):
        return "historical_duplicate"
    if not brief.is_allowed_topic():
        return "topic_not_recent_completed_or_classic"
    if is_excluded_case(brief):
        return "excluded_term"
    if not has_explicit_parties(brief):
        return "missing_explicit_parties"
    if not is_report_completed_candidate(brief):
        return "missing_completed_evidence"
    if not has_usable_source_url(brief.source_url):
        return "missing_usable_source_url"
    evidence_text = "\n".join([brief.deal_value, brief.why, brief.source_title, brief.financial_highlights])
    if not has_deal_value_signal(evidence_text) and not (
        is_authoritative_source_url(brief.source_url)
        and (has_detail_rich_disclosure_signal(brief) or (brief.is_completed and has_rich_disclosure_signal(brief)))
    ):
        return "missing_deal_terms_or_rich_disclosure"
    return "accepted"


def infer_parties_from_name(case_name: str) -> tuple[str, str]:
    parts = re.split(r"收购|并购|入主|吸收合并|私有化|合并|出售|控股|取得", case_name or "", maxsplit=1)
    if len(parts) >= 2:
        return parts[0].strip(), parts[1].strip()
    return "", ""


def is_vague_party(value: str) -> bool:
    text = (value or "").strip()
    if len(text) < 2:
        return True
    compact = re.sub(r"[^A-Za-z0-9\u4e00-\u9fa5]", "", text).lower()
    if compact in {"st", "xst", "sst"}:
        return True
    return any(term in text for term in VAGUE_PARTY_TERMS)


def normalize_identity_part(value: str | None) -> str:
    text = (value or "").lower().strip()
    text = re.sub(r"（.*?）|\(.*?\)", "", text)
    text = PARTY_SUFFIX_RE.sub("", text)
    text = re.sub(r"[\s·・,，.。:：;；/\\|&＋+_-]+", "", text)
    if not text or is_vague_party(text):
        return ""
    return text


def identity_part_matches(left: str, right: str) -> bool:
    left = normalize_identity_part(left)
    right = normalize_identity_part(right)
    if not left or not right:
        return False
    if left == right:
        return True
    if len(left) >= 3 and len(right) >= 3 and (left in right or right in left):
        return True
    return False


def normalize_case_name_for_identity(value: str | None) -> str:
    text = (value or "").lower().strip()
    text = re.sub(r"^(?:weekly|backfill)_\d{8}_\d{6}_", "", text)
    text = re.split(r"[:：]", text, maxsplit=1)[0]
    text = re.sub(r"20\d{2}[年/-]?\d{0,2}[月/-]?\d{0,2}日?", "", text)
    text = re.sub(r"\d+(?:,\d{3})*(?:\.\d+)?(?:亿元|亿美元|亿港元|万元|美元|港元|元|%|％|股|股份|股权)?", "", text)
    text = re.sub(r"(交易复盘|案例分析|并购案例|交易启示|交易观察|并购启示|案例研究|事实复盘)", "", text)
    text = re.sub(r"\W+", "", text)
    return text[:80]


def case_identity_key(brief: CaseBrief) -> str:
    case_name = re.sub(r"^(?:weekly|backfill)_\d{8}_\d{6}_", "", brief.case_name or "")
    case_name_core = re.split(r"[:：]", case_name, maxsplit=1)[0]
    inferred_a, inferred_t = infer_parties_from_name(case_name_core)
    acquirer = normalize_identity_part(brief.acquirer or inferred_a)
    target = normalize_identity_part(brief.target or inferred_t)
    if acquirer and target:
        return f"party:{acquirer}->{target}"
    name_key = normalize_case_name_for_identity(case_name_core)
    return f"name:{name_key}" if name_key else ""


def case_identity_keys(brief: CaseBrief) -> set[str]:
    case_name = re.sub(r"^(?:weekly|backfill)_\d{8}_\d{6}_", "", brief.case_name or "")
    case_name_core = re.split(r"[:：]", case_name, maxsplit=1)[0]
    inferred_a, inferred_t = infer_parties_from_name(case_name_core)
    acquirer = normalize_identity_part(brief.acquirer or inferred_a)
    target = normalize_identity_part(brief.target or inferred_t)
    keys: set[str] = set()
    if acquirer and target:
        keys.add(f"party:{acquirer}->{target}")
    name_key = normalize_case_name_for_identity(case_name_core)
    if name_key:
        keys.add(f"name:{name_key}")
    return keys


def case_key_matches(left: str, right: str) -> bool:
    if not left or not right:
        return False
    if left == right:
        return True
    left_kind, _, left_value = left.partition(":")
    right_kind, _, right_value = right.partition(":")
    if left_kind != right_kind:
        return False
    if left_kind == "party":
        left_a, _, left_t = left_value.partition("->")
        right_a, _, right_t = right_value.partition("->")
        return identity_part_matches(left_a, right_a) and identity_part_matches(left_t, right_t)
    if left_kind == "name":
        return len(left_value) >= 6 and len(right_value) >= 6 and (left_value in right_value or right_value in left_value)
    return False


def key_in_history(key: str, historical_keys: set[str]) -> bool:
    return any(case_key_matches(key, old_key) for old_key in historical_keys)


def any_key_in_history(keys: set[str], historical_keys: set[str]) -> bool:
    return any(key_in_history(key, historical_keys) for key in keys)


def keys_overlap(left: set[str], right: set[str]) -> bool:
    return any(case_key_matches(left_key, right_key) for left_key in left for right_key in right)


def has_explicit_parties(brief: CaseBrief) -> bool:
    acquirer = brief.acquirer
    target = brief.target
    inferred_a, inferred_t = infer_parties_from_name(brief.case_name)
    acquirer = acquirer or inferred_a
    target = target or inferred_t
    if is_vague_party(acquirer) or is_vague_party(target):
        return False
    combined = "\n".join([brief.case_name, brief.acquirer, brief.target])
    return not any(term in combined for term in VAGUE_PARTY_TERMS)


def is_excluded_case(brief: CaseBrief) -> bool:
    terms = excluded_terms()
    if not terms:
        return False
    text = "\n".join([
        brief.case_name,
        brief.acquirer,
        brief.target,
        brief.source_title,
        brief.why,
    ]).lower()
    return any(term in text for term in terms)


def safe_category(value: str | None) -> str:
    if value in CATEGORIES:
        return value
    text = value or ""
    for category in CATEGORIES:
        if category in text:
            return category
    return "依托上市平台持续整合同类资产"


def is_domestic_region(region: str) -> bool:
    text = region or ""
    return any(token in text for token in ("中国", "A股", "港股", "境内", "香港", "中概"))


def infer_completed_year(*values: str) -> str:
    text = " ".join(v or "" for v in values)
    years = re.findall(r"20(?:2[0-6]|1[0-9])", text)
    for year in ("2026", "2025"):
        if year in years:
            return year
    return years[0] if years else ""


def infer_is_completed(*values: str) -> bool:
    text = " ".join(v or "" for v in values)
    return has_completed_signal(text)


def existing_counts(report_root: Path) -> Counter[str]:
    counts: Counter[str] = Counter()
    for category in CATEGORIES:
        counts[category] = 0
    if not report_root.exists():
        return counts
    for path in report_root.rglob("*.docx"):
        for category in CATEGORIES:
            if category in str(path):
                counts[category] += 1
    return counts


def historical_case_keys(report_root: Path) -> set[str]:
    keys: set[str] = set()
    if not report_root.exists():
        return keys
    for manifest in report_root.glob("_manifests/*.json"):
        if manifest.name.startswith("docx_format_validation_"):
            continue
        try:
            data = json.loads(manifest.read_text(encoding="utf-8"))
        except Exception:  # noqa: BLE001
            continue
        rows = data if isinstance(data, list) else []
        for row in rows:
            if not isinstance(row, dict):
                continue
            case_name = str(row.get("case_name") or "").strip()
            if not case_name:
                continue
            inferred_a, inferred_t = infer_parties_from_name(case_name)
            brief = CaseBrief(
                case_name=case_name,
                category=safe_category(row.get("category")),
                region=str(row.get("region") or ""),
                acquirer=str(row.get("acquirer") or inferred_a),
                target=str(row.get("target") or inferred_t),
            )
            keys.update(case_identity_keys(brief))
    for path in report_root.rglob("*.docx"):
        stem = re.sub(r"^(?:weekly|backfill)_\d{8}_\d{6}_", "", path.stem)
        stem = re.split(r"[:：]", stem, maxsplit=1)[0]
        inferred_a, inferred_t = infer_parties_from_name(stem)
        keys.update(case_identity_keys(CaseBrief(case_name=stem, category="", region="", acquirer=inferred_a, target=inferred_t)))
    return keys


def rows_to_briefs(rows: list[dict[str, str]], *, default_classic: bool = False) -> list[CaseBrief]:
    briefs: list[CaseBrief] = []
    for row in rows:
        case_name = str(row.get("case_name") or "").strip()
        if not case_name:
            continue
        inferred_a, inferred_t = infer_parties_from_name(case_name)
        region = str(row.get("region") or "中国")
        completed_year = str(row.get("completed_year") or infer_completed_year(case_name, str(row.get("why") or "")))
        inferred_completed = infer_is_completed(
            case_name,
            str(row.get("source_title") or ""),
            str(row.get("why") or ""),
            str(row.get("deal_status") or ""),
        )
        is_completed = parse_bool(row.get("is_completed"), default=inferred_completed)
        is_classic_raw = row.get("is_classic")
        if is_classic_raw is None:
            is_classic = default_classic or completed_year not in {"2025", "2026"}
        else:
            is_classic = str(is_classic_raw).lower() == "true"
        source_url = unwrap_news_url(str(row.get("source_url") or ""))
        if not is_usable_article_url(source_url):
            source_url = ""
        brief = CaseBrief(
            case_name=case_name,
            category=safe_category(row.get("category")),
            region=region,
            source_title=str(row.get("source_title") or case_name),
            source_url=source_url,
            published_at=str(row.get("published_at") or ""),
            why=str(row.get("why") or "经典或代表性并购案例。"),
            is_domestic=is_domestic_region(region),
            is_classic=is_classic,
            completed_year=completed_year,
            is_completed=is_completed,
            acquirer=str(row.get("acquirer") or inferred_a),
            target=str(row.get("target") or inferred_t),
            deal_value=str(row.get("deal_value") or ""),
            deal_status=str(row.get("deal_status") or ""),
            buyer_motivation=str(row.get("buyer_motivation") or ""),
            seller_motivation=str(row.get("seller_motivation") or ""),
            financial_highlights=str(row.get("financial_highlights") or ""),
        )
        if brief.is_allowed_topic() and has_explicit_parties(brief) and not is_excluded_case(brief):
            briefs.append(brief)
    return briefs


def seed_briefs() -> list[CaseBrief]:
    return rows_to_briefs(CLASSIC_CASE_SEEDS, default_classic=True)


def extended_pool_briefs() -> list[CaseBrief]:
    return rows_to_briefs(EXTENDED_CASE_POOL, default_classic=False)


def candidates_from_weekly(days: int, max_items: int) -> list[RawItem]:
    end = datetime.now(BEIJING_TZ).replace(microsecond=0)
    start = end - timedelta(days=days)
    raw, _errors = fetch_all_candidates(start, end, max_items=max_items)
    return raw


def tavily_report_queries() -> list[str]:
    queries: list[str] = []
    extra = os.getenv("REPORT_TAVILY_EXTRA_QUERIES", "")
    if extra:
        queries.extend(query.strip() for query in re.split(r"[\n]+", extra) if query.strip())
    queries.extend(TAVILY_COMPLETION_QUERIES)
    queries.extend(CASE_DISCOVERY_QUERIES)
    seen: set[str] = set()
    out: list[str] = []
    for query in queries:
        key = query.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(query)
    return out


def fetch_tavily_report_candidates(start: datetime, end: datetime, max_items: int) -> list[RawItem]:
    api_key = os.getenv("TAVILY_API_KEY", "").strip()
    if not api_key:
        LOGGER.info("Tavily report discovery skipped because TAVILY_API_KEY is not configured")
        source_notice("tavily_report_discovery_skipped reason=missing_api_key")
        return []
    if not env_flag("REPORT_ENABLE_TAVILY", default=True):
        LOGGER.info("Tavily report discovery disabled by REPORT_ENABLE_TAVILY")
        source_notice("tavily_report_discovery_skipped reason=disabled")
        return []

    try:
        import requests
    except Exception as exc:  # noqa: BLE001
        LOGGER.warning("Tavily report discovery skipped because requests import failed: %s", exc)
        source_notice(f"tavily_report_discovery_skipped reason=requests_import_failed error={str(exc)[:200]}")
        return []

    endpoint = os.getenv("TAVILY_SEARCH_URL", "https://api.tavily.com/search").strip() or "https://api.tavily.com/search"
    max_queries = env_int("REPORT_TAVILY_MAX_QUERIES", 10, minimum=0)
    results_per_query = min(env_int("REPORT_TAVILY_RESULTS_PER_QUERY", 8, minimum=1), 20)
    timeout = env_int("REPORT_TAVILY_TIMEOUT_SECONDS", 20, minimum=5)
    search_depth = os.getenv("REPORT_TAVILY_SEARCH_DEPTH", "basic").strip() or "basic"
    topic = os.getenv("REPORT_TAVILY_TOPIC", "general").strip() or "general"
    include_raw = env_flag("REPORT_TAVILY_INCLUDE_RAW_CONTENT", default=False)

    raw: list[RawItem] = []
    seen: set[str] = set()
    queries = tavily_report_queries()[:max_queries]
    source_notice(
        f"tavily_report_discovery_start queries={len(queries)} results_per_query={results_per_query} "
        f"depth={search_depth} topic={topic} include_raw={include_raw}"
    )
    for query in queries:
        payload = {
            "query": query,
            "search_depth": search_depth,
            "max_results": results_per_query,
            "topic": topic,
            "start_date": start.strftime("%Y-%m-%d"),
            "end_date": end.strftime("%Y-%m-%d"),
            "include_answer": False,
            "include_raw_content": "text" if include_raw else False,
            "include_images": False,
            "include_image_descriptions": False,
            "include_favicon": False,
            "auto_parameters": False,
            "exact_match": False,
            "include_usage": True,
            "safe_search": False,
        }
        try:
            response = requests.post(
                endpoint,
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json=payload,
                timeout=timeout,
            )
            if response.status_code >= 400:
                LOGGER.warning("Tavily report discovery query failed: status=%s query=%s body=%s", response.status_code, query, response.text[:300])
                source_notice(f"tavily_report_query_failed status={response.status_code} query={query[:160]}")
                continue
            data = response.json()
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("Tavily report discovery query failed: query=%s error=%s", query, exc)
            source_notice(f"tavily_report_query_failed query={query[:160]} error={str(exc)[:220]}")
            continue
        accepted = 0
        for result in data.get("results") or []:
            if not isinstance(result, dict):
                continue
            title = clean_cell(result.get("title"))
            url = unwrap_news_url(clean_cell(result.get("url")))
            if not title or not is_usable_article_url(url):
                continue
            content = clean_cell(result.get("content"))
            raw_content = clean_cell(result.get("raw_content")) if include_raw else ""
            summary = " ".join(x for x in [content, raw_content[:1200]] if x).strip()
            item = RawItem(
                title=title,
                url=url,
                source_name="Tavily Search",
                source_url="https://www.tavily.com/",
                published_at=clean_cell(result.get("published_date") or result.get("published_at") or ""),
                summary=summary,
                region_hint="中国/全球",
                query=query,
            )
            key = item.stable_key()
            if key in seen:
                continue
            seen.add(key)
            raw.append(item)
            accepted += 1
            if len(raw) >= max_items:
                LOGGER.info("Tavily report discovery reached cap=%s", max_items)
                source_notice(f"tavily_report_discovery_cap reached={len(raw)} cap={max_items}")
                return sorted(raw, key=raw_report_item_score, reverse=True)
        usage = data.get("usage") if isinstance(data, dict) else None
        LOGGER.info("Tavily report discovery query done: query=%s accepted=%s usage=%s", query, accepted, usage or "-")
        source_notice(f"tavily_report_query_done accepted={accepted} total={len(raw)} query={query[:160]}")
    LOGGER.info("Tavily report discovery collected=%s queries=%s", len(raw), len(queries))
    source_notice(f"tavily_report_discovery_done collected={len(raw)} queries={len(queries)}")
    return sorted(raw, key=raw_report_item_score, reverse=True)


def lightweight_weekly_candidates(days: int, max_items: int) -> list[RawItem]:
    end = datetime.now(BEIJING_TZ).replace(microsecond=0)
    start = end - timedelta(days=days)
    from mna_weekly_tracker.sources_rich import fetch_bing_news, fetch_google_news

    raw: list[RawItem] = []
    seen: set[str] = set()
    for query in CASE_DISCOVERY_QUERIES:
        items: list[RawItem] = []
        items.extend(fetch_google_news(query, start, end, source_name="Google News - report live", source_url="https://news.google.com/", region_hint="中国"))
        items.extend(fetch_bing_news(query, start, end, source_name="Bing News - report live", source_url="https://www.bing.com/news/search", region_hint="中国"))
        for item in items:
            key = item.stable_key()
            if key in seen:
                continue
            seen.add(key)
            raw.append(item)
    tavily_items = fetch_tavily_report_candidates(start, end, max_items=max_items)
    for item in tavily_items:
        key = item.stable_key()
        if key in seen:
            continue
        seen.add(key)
        raw.append(item)
    raw = sorted(raw, key=raw_report_item_score, reverse=True)
    LOGGER.info("Lightweight weekly report candidates collected=%s tavily=%s returning=%s", len(raw), len(tavily_items), min(len(raw), max_items))
    source_notice(f"lightweight_report_discovery_done collected={len(raw)} tavily={len(tavily_items)} returning={min(len(raw), max_items)}")
    return raw[:max_items]


def latest_weekly_workbook(output_dir: Path) -> Path | None:
    paths = sorted(output_dir.glob("并购案例一览_*.xlsx"))
    return paths[-1] if paths else None


def recent_weekly_workbooks(output_dir: Path, limit: int | None = None) -> list[Path]:
    paths = sorted(output_dir.glob("并购案例一览_*.xlsx"), reverse=True)
    if limit is None:
        limit = env_int("REPORT_WEEKLY_WORKBOOK_LOOKBACK", 12, minimum=1)
    return paths[: max(limit, 1)]


def _excel_bool_completed(status: str, remark: str, intro: str) -> bool:
    return has_completed_signal(" ".join([status, remark, intro]))


def _excel_completed_year(*values: str) -> str:
    year = infer_completed_year(*values)
    return year or str(datetime.now(BEIJING_TZ).year)


def _brief_from_weekly_excel_row(row: dict[str, object]) -> CaseBrief | None:
    acquirer = clean_cell(row.get("并购方"))
    target = clean_cell(row.get("目标方"))
    if is_vague_party(acquirer) or is_vague_party(target):
        return None
    category = safe_category(clean_cell(row.get("案例分类")))
    intro = clean_cell(row.get("案例一句话简介"))
    deal_time = clean_cell(row.get("交易时间"))
    deal_value = clean_cell(row.get("交易对价"))
    deal_status = clean_cell(row.get("交易状态"))
    remark = clean_cell(row.get("备注"))
    source_name = clean_cell(row.get("来源名称"))
    source_url = unwrap_news_url(clean_cell(row.get("URL")))
    if not is_usable_article_url(source_url):
        source_url = ""
    case_name = f"{acquirer}收购{target}"
    if "SPAC" in category or any(token in intro.lower() for token in ("spac", "de-spac", "despac")):
        case_name = f"{target}通过SPAC合并上市"
    elif any(token in intro for token in ("出售", "剥离", "转让")):
        case_name = f"{acquirer}受让{target}"
    elif "吸收合并" in intro:
        case_name = f"{acquirer}吸收合并{target}"
    evidence = "；".join(x for x in [intro, remark] if x and not is_placeholder(x))
    transaction_terms = extract_transaction_terms(deal_value, intro, remark)
    return CaseBrief(
        case_name=case_name,
        category=category,
        region=clean_cell(row.get("地区")) or "中国",
        source_title=intro or source_name or case_name,
        source_url=source_url,
        published_at=clean_cell(row.get("发布日期")),
        why=evidence or intro,
        is_domestic=is_domestic_region(clean_cell(row.get("地区"))),
        is_classic=False,
        completed_year=_excel_completed_year(deal_time, deal_status, intro),
        is_completed=_excel_bool_completed(deal_status, remark, intro),
        acquirer=acquirer,
        target=target,
        deal_value=transaction_terms or ("" if is_placeholder(deal_value) else deal_value),
        deal_status="；".join(x for x in [deal_time, deal_status] if x and not is_placeholder(x)),
        buyer_motivation=evidence,
        seller_motivation=remark if remark and not is_placeholder(remark) else "",
        financial_highlights=evidence if has_deal_number(evidence) else transaction_terms,
    )


def briefs_from_latest_weekly_workbook(output_dir: Path) -> list[CaseBrief]:
    workbook_paths = recent_weekly_workbooks(output_dir)
    if not workbook_paths:
        LOGGER.info("No weekly Excel workbook found under %s", output_dir)
        return []
    briefs: list[CaseBrief] = []
    seen_keys: set[str] = set()
    for workbook_path in workbook_paths:
        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            sheet = workbook["周度并购案例"] if "周度并购案例" in workbook.sheetnames else workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [clean_cell(value) for value in next(rows, [])]
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("Failed to read weekly Excel workbook for report candidates: %s error=%s", workbook_path, exc)
            continue
        before = len(briefs)
        for values in rows:
            row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
            brief = _brief_from_weekly_excel_row(row)
            if not brief or not brief.is_allowed_topic() or not has_explicit_parties(brief) or is_excluded_case(brief):
                continue
            keys = case_identity_keys(brief) or {brief.key()}
            if keys_overlap(keys, seen_keys):
                continue
            seen_keys.update(keys)
            briefs.append(brief)
        LOGGER.info("Loaded %s report candidates from weekly Excel: %s", len(briefs) - before, workbook_path)
    LOGGER.info("Loaded %s report candidates from %s recent weekly Excel workbooks", len(briefs), len(workbook_paths))
    return briefs


def raw_items_from_latest_weekly_workbook(output_dir: Path, max_items: int = 220) -> list[RawItem]:
    workbook_paths = recent_weekly_workbooks(output_dir)
    if not workbook_paths:
        return []
    raw_items: list[RawItem] = []
    seen: set[str] = set()
    for workbook_path in workbook_paths:
        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            if "原始候选" not in workbook.sheetnames:
                continue
            sheet = workbook["原始候选"]
            rows = sheet.iter_rows(values_only=True)
            headers = [clean_cell(value) for value in next(rows, [])]
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("Failed to read raw candidates from weekly Excel workbook: %s error=%s", workbook_path, exc)
            continue
        for values in rows:
            row = {headers[index]: values[index] for index in range(min(len(headers), len(values)))}
            title = clean_cell(row.get("标题"))
            url = unwrap_news_url(clean_cell(row.get("URL")))
            if not title or not is_usable_article_url(url):
                continue
            item = RawItem(
                title=title,
                url=url,
                source_name=clean_cell(row.get("来源名称")),
                source_url="",
                published_at=clean_cell(row.get("发布时间")),
                summary=clean_cell(row.get("摘要")),
                region_hint=clean_cell(row.get("地区")) or "中国",
                query=clean_cell(row.get("查询词")),
            )
            key = item.stable_key()
            if key in seen:
                continue
            seen.add(key)
            raw_items.append(item)
    raw_items = sorted(raw_items, key=raw_report_item_score, reverse=True)
    if len(raw_items) > max_items:
        raw_items = raw_items[:max_items]
    LOGGER.info("Loaded %s prioritized raw report candidates from %s recent weekly Excel workbooks", len(raw_items), len(workbook_paths))
    return raw_items


def _official_pdf_text(url: str, *, max_pages: int = 4, max_chars: int = 9000) -> str:
    if not is_authoritative_source_url(url) or not url.lower().split("?", 1)[0].endswith(".pdf"):
        return ""
    try:
        import requests
        from pypdf import PdfReader

        response = requests.get(
            url,
            headers={"User-Agent": "Mozilla/5.0 AppleWebKit/537.36 Chrome/124 Safari/537.36"},
            timeout=25,
        )
        response.raise_for_status()
        content = response.content[: 12 * 1024 * 1024]
        if not content.startswith(b"%PDF") and b"%PDF" not in content[:1024]:
            return ""
        reader = PdfReader(BytesIO(content))
        pages: list[str] = []
        for page in reader.pages[:max_pages]:
            pages.append(page.extract_text() or "")
        return clean_cell("\n".join(pages))[:max_chars]
    except Exception as exc:  # noqa: BLE001
        LOGGER.debug("Failed to extract official PDF text for candidate: url=%s error=%s", url, exc)
        return ""


def _extract_stock_identity(text: str, summary: str) -> tuple[str, str]:
    short = ""
    full = ""
    short_match = re.search(r"证券简称[:：]\s*([^\s，。；;]+)", f"{summary} {text[:800]}")
    if short_match:
        short = clean_cell(short_match.group(1))
    head = text[:1200]
    prefix_match = re.match(r"\s*([A-Za-z0-9\u4e00-\u9fa5（）()*ST\s]{4,80}(?:股份有限公司|有限责任公司|有限公司))\s+关于", head, re.I)
    if prefix_match:
        full = clean_cell(prefix_match.group(1))
    full_match = re.search(
        r"证券代码[:：]\s*[0-9A-Za-z]+.*?证券简称[:：]\s*[^\s，。；;]+.*?([A-Za-z0-9\u4e00-\u9fa5（）()]{4,80}(?:股份有限公司|有限责任公司|有限公司))\s+关于",
        head,
        re.S,
    )
    if not full and full_match:
        full = clean_cell(full_match.group(1))
    if not full:
        full_match = re.search(r"([A-Za-z0-9\u4e00-\u9fa5（）()]{4,80}(?:股份有限公司|有限责任公司|有限公司))（以下简称[“\"]?(?:公司|[^”\"]{2,12})", head)
        if full_match:
            full = clean_cell(full_match.group(1))
    if any(token in full for token in ("协议转让", "过户登记", "通过协议")):
        full = ""
    return full, short


def _alias_map(text: str) -> dict[str, str]:
    aliases: dict[str, str] = {}
    for match in re.finditer(
        r"([A-Za-z0-9\u4e00-\u9fa5（）()\s]{4,90}(?:股份有限公司|有限责任公司|有限公司|合伙企业（有限合伙）|合伙企业\(有限合伙\)|基金))（以下简称[“\"]([^”\"]{2,20})[”\"]",
        text[:5000],
    ):
        aliases[clean_cell(match.group(2))] = _clean_party_name(match.group(1))
    return aliases


def _clean_party_name(value: str) -> str:
    value = clean_cell(value)
    value = re.sub(r"^(?:[一二三四五六七八九十]+[、.．]\s*)?(?:交易基本情况|本次交易的基本情况|股份过户登记的情况|其他说明)", "", value).strip()
    for marker in ("方式向", "的境内全资子公司", "境内全资子公司", "全资子公司", "向"):
        if marker in value:
            value = value.split(marker)[-1]
    value = re.split(r"（以下简称|以下简称|（代表|代表|及其一致行动人|承诺|于20\d{2}|于\s*20\d{2}|并|，|。|；|;|、", value, maxsplit=1)[0]
    value = value.strip(" “”\"'（）()：:")
    if "（有限合伙" in value and "）" not in value.split("（有限合伙", 1)[-1]:
        value = f"{value}）"
    value = re.sub(r"^[（(]?\d+[）)]", "", value).strip()
    value = re.sub(r"^(?:与|和|同|公司|本公司|控股股东|实际控制人|一致行动人|自然人)", "", value).strip()
    value = re.sub(r"(?:共同|进一步)$", "", value).strip()
    if re.search(r"[\u4e00-\u9fa5]", value):
        value = re.sub(r"\s+", "", value)
    return value


def _resolve_alias(name: str, aliases: dict[str, str]) -> str:
    cleaned = _clean_party_name(name)
    return aliases.get(cleaned, cleaned)


def _extract_buyer_from_transfer(text: str, aliases: dict[str, str]) -> str:
    text = clean_cell(text)
    patterns = (
        r"(?:公司控股股东|公司的控股股东|控股股东)(?:由[^，。；]{1,80})?变更为([A-Za-z0-9\u4e00-\u9fa5（）()·\s]{2,100}?)(?:，|。|；|,|;|$)",
        r"([A-Za-z0-9\u4e00-\u9fa5（）()·\s]{2,100}?)(?:将成为|成为)(?:上市公司|公司)?控股股东",
        r"(?:转让方|控股股东|实际控制人|股东)[^。；]{0,120}?向([A-Za-z0-9\u4e00-\u9fa5（）()·\s、及其一致行动人]{2,140}?)(?:转让|协议转让)",
        r"转让给([A-Za-z0-9\u4e00-\u9fa5（）()·\s、及其一致行动人]{2,120}?)(?:，|。|；|,|;|股份|$)",
        r"受让方[）)]?\s*(?:为|系|：|:)\s*([A-Za-z0-9\u4e00-\u9fa5（）()·\s]{2,100}?)(?:，|。|；|,|;|$)",
        r"与([A-Za-z0-9\u4e00-\u9fa5（）()·\s]{2,120}?)(?:签订|签署).*?(?:股份转让|股权转让)",
        r"(?:公司实际控制人|实际控制人)[^。；]{0,60}变更为([A-Za-z0-9\u4e00-\u9fa5（）()·\s]{2,100}?)(?:，|。|；|,|;|$)",
    )
    for pattern in patterns:
        match = re.search(pattern, text[:5000])
        if match:
            buyer = _resolve_alias(match.group(1), aliases)
            if (
                buyer
                and not is_vague_party(buyer)
                and not re.search(r"^(?:其|本次|上述|公司|上市公司|在)", buyer)
                and not re.search(r"(?:持有公司|股份协议|过户登记|转让事项|受让方|转让方|证券过户)", buyer)
            ):
                return buyer
    return ""


def _sentences_with(text: str, tokens: tuple[str, ...], *, limit: int = 4) -> str:
    parts = re.split(r"(?<=[。；;])", text)
    chosen: list[str] = []
    for part in parts:
        cleaned = clean_cell(part)
        if cleaned and any(token in cleaned for token in tokens):
            chosen.append(cleaned)
        if len(chosen) >= limit:
            break
    return "；".join(chosen)[:900]


def _completion_date_status(text: str, published_at: str) -> str:
    for pattern in (
        r"过户日期为\s*(20\d{2}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日)",
        r"股份过户日期为\s*(20\d{2}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日)",
        r"(20\d{2}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日)[^。；]{0,30}(?:完成|办理了|取得).*?(?:过户|交割|工商变更)",
        r"(20\d{2}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日)[^。；]{0,30}(?:过户登记|股权交割|工商变更登记)",
    ):
        match = re.search(pattern, text[:5000])
        if match:
            return f"{clean_cell(match.group(1))}；已完成过户/交割"
    date = (published_at or "").split("T", 1)[0]
    return f"{date}；已完成过户/交割" if date else "已完成过户/交割"


def _percent_values(text: str) -> list[float]:
    values: list[float] = []
    for match in re.finditer(r"(\d+(?:\.\d+)?)\s*%", text):
        try:
            values.append(float(match.group(1)))
        except ValueError:
            continue
    return values


def _largest_percent_label(text: str) -> str:
    values = _percent_values(text)
    if not values:
        return ""
    value = max(values)
    label = f"{value:g}%"
    return label


def _transfer_percent_label(text: str) -> str:
    text = clean_cell(text)
    patterns = (
        r"(?:转让|受让)[^。；]{0,120}?占(?:公司|上市公司|总股本|公司总股本|上市公司股份总数)?[^。；]{0,30}?(\d+(?:\.\d+)?)\s*%",
        r"(?:过户数量|过户股份数量|合计过户数量)[^。；]{0,100}?占(?:公司|上市公司|总股本|公司总股本)?[^。；]{0,30}?(\d+(?:\.\d+)?)\s*%",
        r"(?:本次协议转让|本次股份转让|本次权益变动)[^。；]{0,160}?(\d+(?:\.\d+)?)\s*%",
        r"(\d+(?:\.\d+)?)\s*%[^。；]{0,40}(?:股份|股权)[^。；]{0,40}(?:转让|过户|受让)",
    )
    for pattern in patterns:
        match = re.search(pattern, text[:7000])
        if match:
            try:
                return f"{float(match.group(1)):g}%"
            except ValueError:
                continue
    return ""


def _brief_from_completed_asset_acquisition(item: RawItem, text: str, company_full: str, company_short: str, aliases: dict[str, str]) -> CaseBrief | None:
    head = text[:6000]
    if not re.search(r"(?:购买|收购|重大资产购买|发行股份购买资产)", " ".join([item.title, head]), re.I):
        return None
    if re.search(r"(?:减值测试|业绩承诺期满|独立财务顾问|法律意见书|核查意见)", item.title):
        return None
    if not re.search(r"(?:标的资产.*?(?:过户|完成)|股权交割|工商变更登记手续|完成交割|完成过户|交易完成)", head):
        return None
    target = ""
    details_match = re.search(r"具体包括[:：]\s*(.{20,260}?)(?:。|根据|二、)", head)
    if details_match:
        stakes = re.findall(r"([A-Za-z0-9\u4e00-\u9fa5（）()·\s]{2,80}?有限公司)\s*(\d+(?:\.\d+)?\s*%股权)", details_match.group(1))
        if stakes:
            target = "、".join(f"{_clean_party_name(name)}{clean_cell(stake)}" for name, stake in stakes[:3])
    if not target:
        patterns = (
            r"(?:购买|收购).*?持有的([A-Za-z0-9\u4e00-\u9fa5（）()·\s]{2,90}?有限公司).*?(\d+(?:\.\d+)?\s*%股权)",
            r"(?:购买|收购)([A-Za-z0-9\u4e00-\u9fa5（）()·\s]{2,90}?有限公司).*?(\d+(?:\.\d+)?\s*%股权)",
            r"(?:购买|收购).*?持有的([^。；]{2,90}?)(?:（以下简称[“\"]标的公司[”\"]）)?\s*(\d+(?:\.\d+)?\s*%股权)",
        )
        for pattern in patterns:
            target_match = re.search(pattern, head, re.I)
            if not target_match:
                continue
            raw_target = _resolve_alias(target_match.group(1), aliases)
            stake = clean_cell(target_match.group(2))
            target = f"{raw_target}{stake}" if stake and stake not in raw_target else raw_target
            break
    if not target and "标的公司" in aliases:
        stake_match = re.search(r"标的公司\s*(\d+(?:\.\d+)?\s*%股权)", head)
        stake = clean_cell(stake_match.group(1)) if stake_match else ""
        target = f"{aliases['标的公司']}{stake}"
    if not target or is_vague_party(target):
        return None
    acquirer = company_full or company_short
    if not acquirer or is_vague_party(acquirer):
        return None
    deal_text = _sentences_with(text, ("标的资产", "购买资产", "股权", "亿元", "万元", "美元", "股权交割", "工商变更登记手续", "完成", "过户", "发行股份"), limit=7)
    terms = extract_transaction_terms(deal_text, head)
    brief = CaseBrief(
        case_name=f"{company_short or acquirer}收购{target}",
        category="依托上市平台持续整合同类资产",
        region=item.region_hint or "中国",
        source_title=item.title,
        source_url=unwrap_news_url(item.url),
        published_at=item.published_at,
        why=deal_text or item.title,
        is_domestic=True,
        is_classic=False,
        completed_year=_excel_completed_year(item.published_at, text[:1000]),
        is_completed=True,
        acquirer=acquirer,
        target=target,
        deal_value=terms,
        deal_status=_completion_date_status(text, item.published_at),
        buyer_motivation=deal_text,
        seller_motivation=deal_text,
        financial_highlights=deal_text if has_deal_number(deal_text) else terms,
    )
    return brief if is_report_ready_candidate(brief) else None


def _brief_from_completed_control_transfer(item: RawItem, text: str, company_full: str, company_short: str, aliases: dict[str, str]) -> CaseBrief | None:
    head = text[:7000]
    if not any(hint in head for hint in CONTROL_COMPLETION_HINTS):
        if not re.search(r"(?:控股股东|实际控制人)[^。；]{0,50}(?:变更为|发生变更)", head):
            return None
    if any(hint in head[:3000] for hint in NON_CONTROL_HINTS):
        return None
    buyer = _extract_buyer_from_transfer(text, aliases)
    target = company_full or company_short
    if not buyer or not target or is_vague_party(buyer) or is_vague_party(target):
        return None
    deal_text = _sentences_with(text, ("转让", "过户", "控股股东", "实际控制人", "控制权", "表决权", "股", "%", "万元", "亿元", "元/股"), limit=8)
    terms = extract_transaction_terms(deal_text, text[:3000])
    brief = CaseBrief(
        case_name=f"{buyer}受让{company_short or target}控制权股份",
        category="上市公司控股权并购",
        region=item.region_hint or "中国",
        source_title=item.title,
        source_url=unwrap_news_url(item.url),
        published_at=item.published_at,
        why=deal_text or item.title,
        is_domestic=True,
        is_classic=False,
        completed_year=_excel_completed_year(item.published_at, text[:1000]),
        is_completed=True,
        acquirer=buyer,
        target=target,
        deal_value=terms,
        deal_status=_completion_date_status(text, item.published_at),
        buyer_motivation=deal_text,
        seller_motivation=deal_text,
        financial_highlights=deal_text if has_deal_number(deal_text) else terms,
    )
    return brief if is_report_ready_candidate(brief) else None


def _brief_from_completed_asset_swap(item: RawItem, text: str, company_full: str, company_short: str, aliases: dict[str, str]) -> CaseBrief | None:
    head = text[:7000]
    if "资产置换" not in " ".join([item.title, head]) or not re.search(r"(?:完成交割|交割完成|资产交割确认书|工商登记手续均已完成)", head):
        return None
    acquirer = company_full or company_short
    if not acquirer or is_vague_party(acquirer):
        return None
    target = ""
    placed_in = re.search(r"置入的资产为(.{20,900}?)(?:置出资产|具体内容|本次交易已经|二、)", head)
    search_area = placed_in.group(1) if placed_in else head
    listed_stakes: list[str] = []
    if "包含" in search_area:
        contains_text = search_area.split("包含", 1)[-1]
        for segment in re.split(r"[、，,]|和", contains_text):
            if not re.search(r"100\s*%股\s*权", segment):
                continue
            name = re.sub(r"（?以下.{0,20}?简称.*", "", segment)
            name = re.sub(r"100\s*%股\s*权.*", "", name)
            name = re.sub(r"^(?:及|和|包含|持有的)", "", name).strip()
            name = _clean_party_name(name)
            if name and not is_vague_party(name) and name not in listed_stakes:
                listed_stakes.append(name)
    if listed_stakes:
        target = "、".join(f"{name}100%股权" for name in listed_stakes[:4])
    if not target:
        named_stakes = re.findall(
            r"([A-Za-z0-9\u4e00-\u9fa5（）()·]{4,90}?(?:公司|有限公司))（以下简称[“\"][^”\"]{2,20}[”\"]）\s*100\s*%股权",
            search_area,
        )
        if named_stakes:
            target = "、".join(f"{_clean_party_name(name)}100%股权" for name in named_stakes[:4])
    if not target:
        alias_targets = []
        for alias, value in aliases.items():
            if alias in {"公司", "本公司", "上市公司", "标的公司"} or len(alias) < 4:
                continue
            if re.search(rf"{re.escape(alias)}[^。；]{{0,25}}?100\s*%股权", search_area) and value and value not in alias_targets:
                alias_targets.append(value)
        if alias_targets:
            target = "、".join(f"{name}100%股权" for name in alias_targets[:4])
    if not target:
        stakes = re.findall(r"([A-Za-z0-9\u4e00-\u9fa5（）()·]{2,80}?(?:公司|有限公司))[^。；]{0,20}?100\s*%股权", search_area)
        if stakes:
            target = "、".join(f"{_resolve_alias(name, aliases)}100%股权" for name in stakes[:4])
    if not target:
        fallback_alias_targets = [value for key, value in aliases.items() if any(token in key for token in ("公司", "资产", "标的"))]
        if fallback_alias_targets:
            target = "、".join(fallback_alias_targets[:4])
    if not target or is_vague_party(target):
        return None
    deal_text = _sentences_with(text, ("资产置换", "置入", "置出", "100%股权", "完成交割", "工商登记", "合并报表"), limit=8)
    terms = extract_transaction_terms(deal_text, head) or "资产置换；100%股权"
    brief = CaseBrief(
        case_name=f"{company_short or acquirer}置入{target}",
        category="整合一级资产+资本化",
        region=item.region_hint or "中国",
        source_title=item.title,
        source_url=unwrap_news_url(item.url),
        published_at=item.published_at,
        why=deal_text or item.title,
        is_domestic=True,
        is_classic=False,
        completed_year=_excel_completed_year(item.published_at, text[:1000]),
        is_completed=True,
        acquirer=acquirer,
        target=target,
        deal_value=terms,
        deal_status=_completion_date_status(text, item.published_at),
        buyer_motivation=deal_text,
        seller_motivation=deal_text,
        financial_highlights=deal_text if has_deal_number(deal_text) else terms,
    )
    return brief if is_report_ready_candidate(brief) else None


def _brief_from_completed_tender_offer(item: RawItem, text: str, company_full: str, company_short: str, aliases: dict[str, str]) -> CaseBrief | None:
    head = text[:7000]
    title_has_offer = "要约收购" in item.title
    positive_offer = title_has_offer or re.search(r"(?:要约收购结果|要约收购[^。；]{0,80}(?:完成过户|过户完成|完成|期限届满))", head)
    negative_offer = re.search(r"(?:不涉及|不触及|未触及)要约收购", head[:3000])
    if not positive_offer or negative_offer:
        return None
    buyer = ""
    title_match = re.search(r"关于(.{2,100}?)要约收购", item.title)
    if title_match:
        buyer = _resolve_alias(title_match.group(1), aliases)
    if not buyer:
        buyer = _extract_buyer_from_transfer(text, aliases)
    target = company_full or company_short
    if not buyer or not target or is_vague_party(buyer) or is_vague_party(target):
        return None
    deal_text = _sentences_with(text, ("要约收购", "完成过户", "收购结果", "股份", "股", "%", "元/股", "总价"), limit=8)
    terms = extract_transaction_terms(deal_text, head)
    brief = CaseBrief(
        case_name=f"{buyer}要约收购{company_short or target}股份",
        category="上市公司控股权并购",
        region=item.region_hint or "中国",
        source_title=item.title,
        source_url=unwrap_news_url(item.url),
        published_at=item.published_at,
        why=deal_text or item.title,
        is_domestic=True,
        is_classic=False,
        completed_year=_excel_completed_year(item.published_at, text[:1000]),
        is_completed=True,
        acquirer=buyer,
        target=target,
        deal_value=terms,
        deal_status=_completion_date_status(text, item.published_at),
        buyer_motivation=deal_text,
        seller_motivation=deal_text,
        financial_highlights=deal_text if has_deal_number(deal_text) else terms,
    )
    return brief if is_report_ready_candidate(brief) else None


def _brief_from_completed_material_share_transfer(item: RawItem, text: str, company_full: str, company_short: str, aliases: dict[str, str]) -> CaseBrief | None:
    head = text[:7000]
    title_head = " ".join([item.title, head])
    if not re.search(r"(?:协议转让|股份转让).{0,80}(?:完成过户|过户完成|完成过户登记|证券过户登记确认书)", title_head):
        return None
    buyer = _extract_buyer_from_transfer(text, aliases)
    target_company = company_full or company_short
    if not buyer or not target_company or is_vague_party(buyer) or is_vague_party(target_company):
        return None
    transfer_percent = _transfer_percent_label(head)
    transfer_percent_value = 0.0
    if transfer_percent:
        try:
            transfer_percent_value = float(transfer_percent.rstrip("%"))
        except ValueError:
            transfer_percent_value = 0.0
    max_percent = transfer_percent_value or max(_percent_values(head) or [0.0])
    has_control_change = any(hint in head for hint in CONTROL_COMPLETION_HINTS) or re.search(r"(?:控股股东|实际控制人)[^。；]{0,50}(?:变更为|发生变更)", head)
    if max_percent < 5 and not has_control_change:
        return None
    deal_text = _sentences_with(
        text,
        ("协议转让", "股份转让", "过户", "受让", "转让价格", "总价款", "占公司", "持股", "元/股", "%", "成为公司持股"),
        limit=8,
    )
    terms = extract_transaction_terms(deal_text, head)
    percent_label = transfer_percent or _largest_percent_label(head)
    if percent_label and percent_label not in terms:
        terms = "；".join(x for x in [terms, f"{percent_label}股份"] if x)
    category = "上市公司控股权并购" if has_control_change else "上市公司+PE"
    case_target = f"{company_short or target_company}{percent_label}股份" if percent_label else f"{company_short or target_company}股份"
    brief = CaseBrief(
        case_name=f"{buyer}受让{case_target}",
        category=category,
        region=item.region_hint or "中国",
        source_title=item.title,
        source_url=unwrap_news_url(item.url),
        published_at=item.published_at,
        why=deal_text or item.title,
        is_domestic=True,
        is_classic=False,
        completed_year=_excel_completed_year(item.published_at, text[:1000]),
        is_completed=True,
        acquirer=buyer,
        target=case_target,
        deal_value=terms,
        deal_status=_completion_date_status(text, item.published_at),
        buyer_motivation=deal_text,
        seller_motivation=deal_text,
        financial_highlights=deal_text if has_deal_number(deal_text) else terms,
    )
    return brief if is_report_ready_candidate(brief) else None


def completion_briefs_from_raw_items(raw_items: list[RawItem], target_count: int) -> list[CaseBrief]:
    """Deterministically promote official completion announcements into report candidates."""
    briefs: list[CaseBrief] = []
    seen: set[str] = set()
    pdf_budget = env_int("REPORT_RAW_COMPLETION_PDF_MAX", 160, minimum=0)
    scanned = 0
    hinted = 0
    official = 0
    pdf_attempted = 0
    pdf_completed = 0
    promoted_by_parser: Counter[str] = Counter()
    for item in sorted(raw_items, key=raw_report_item_score, reverse=True):
        scanned += 1
        if len(briefs) >= target_count or pdf_budget <= 0:
            break
        title_text = " ".join([item.title, item.summary, item.query])
        if not any(hint in title_text for hint in COMPLETION_ANNOUNCEMENT_HINTS):
            continue
        hinted += 1
        if any(hint in item.title for hint in COMPLETION_SKIP_TITLE_HINTS):
            continue
        url = unwrap_news_url(item.url)
        if not is_authoritative_source_url(url):
            continue
        official += 1
        text = _official_pdf_text(url)
        pdf_budget -= 1
        pdf_attempted += 1
        if not text or not has_completed_signal(" ".join([item.title, text[:1500]])):
            continue
        pdf_completed += 1
        company_full, company_short = _extract_stock_identity(text, item.summary)
        aliases = _alias_map(text)
        candidates = [
            ("asset_acquisition", _brief_from_completed_asset_acquisition(item, text, company_full, company_short, aliases)),
            ("asset_swap", _brief_from_completed_asset_swap(item, text, company_full, company_short, aliases)),
            ("control_transfer", _brief_from_completed_control_transfer(item, text, company_full, company_short, aliases)),
            ("tender_offer", _brief_from_completed_tender_offer(item, text, company_full, company_short, aliases)),
            ("material_share_transfer", _brief_from_completed_material_share_transfer(item, text, company_full, company_short, aliases)),
        ]
        for parser_name, brief in candidates:
            if not brief:
                continue
            keys = case_identity_keys(brief) or {brief.key()}
            if keys_overlap(keys, seen):
                continue
            seen.update(keys)
            briefs.append(brief)
            promoted_by_parser[parser_name] += 1
            LOGGER.info("Promoted completed raw official announcement: parser=%s case=%s", parser_name, brief.case_name)
            break
    LOGGER.info(
        "Completion promotion summary: scanned=%s hinted=%s official=%s pdf_attempted=%s pdf_completed=%s promoted=%s by_parser=%s",
        scanned,
        hinted,
        official,
        pdf_attempted,
        pdf_completed,
        len(briefs),
        dict(promoted_by_parser),
    )
    return briefs


def summarize_raw_items(raw_items: list[RawItem], target_count: int) -> list[CaseBrief]:
    if not raw_items:
        return []
    exclude = "、".join(excluded_terms())
    rows: list[dict[str, str]] = []
    max_chunks = max(1, int(os.getenv("REPORT_RAW_SUMMARY_CHUNKS", "1")))
    chunk_size = max(40, int(os.getenv("REPORT_RAW_SUMMARY_CHUNK_SIZE", "220")))
    max_scan = min(len(raw_items), max_chunks * chunk_size)
    for chunk_index, start in enumerate(range(0, max_scan, chunk_size), start=1):
        sample = [item.as_dict() for item in raw_items[start: min(start + chunk_size, len(raw_items))]]
        if not sample:
            continue
        LOGGER.info("Raw candidate LLM summary chunk start: chunk=%s start=%s size=%s target=%s", chunk_index, start, len(sample), target_count)
        prompt_parts = [
            "从候选新闻/公告中筛选适合写成并购案例分析报告的交易。",
            f"选题规则：{TOPIC_SELECTION_RULES}",
            "只选择已经完成交割、完成合并、完成资产过户或完成股权转让的案例；未完成交易不得输出为Word报告候选。",
            "优先选择并购方和并购标的名称明确、交易金额、估值、股权比例或支付方式线索明确的案例。",
            f"严禁选择并购方或标的名称包含这些模糊词的案例：{'、'.join(VAGUE_PARTY_TERMS)}。",
            "交易主体、交易事件、完成/交割/过户时间、交易对价或股权比例和启示维度要清楚；剔除纯传闻、纯政策、纯市场评论、意向、审批中、进行中、问询阶段、终止交易和交易主体不明案例。",
        ]
        if exclude:
            prompt_parts.append(f"不要选择包含这些主体或关键词的案例：{exclude}。")
        prompt_parts.extend([
            f"最多输出 {max(1, target_count - len(rows))} 个。分类只能用：{json.dumps(CATEGORIES, ensure_ascii=False)}。",
            "source_url必须使用候选里的原始公告/媒体链接，不得使用news.google.com、news.google.com/rss/articles、bing.com/news/apiclick等聚合跳转链接；无法确认原文链接时留空。",
            "输出格式：{\"cases\":[{\"case_name\":...,\"category\":...,\"region\":...,\"source_title\":...,\"source_url\":...,\"published_at\":...,\"why\":...,\"is_domestic\":true/false,\"completed_year\":\"2025或2026等\",\"is_completed\":true/false,\"is_classic\":true/false,\"acquirer\":...,\"target\":...,\"deal_value\":...,\"deal_status\":...,\"buyer_motivation\":...,\"seller_motivation\":...,\"financial_highlights\":...}]}。",
            f"候选：{json.dumps(sample, ensure_ascii=False)}",
        ])
        messages = [
            {"role": "system", "content": "你是并购案例研究选题编辑。只输出 JSON。"},
            {"role": "user", "content": "".join(prompt_parts)},
        ]
        try:
            payload = chat_json(messages)
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("Raw candidate summarization chunk failed; continuing with deterministic/fallback candidates: chunk=%s error=%s", chunk_index, exc)
            continue
        accepted = 0
        rejected = 0
        for row in payload.get("cases", []):
            if not isinstance(row, dict):
                rejected += 1
                continue
            case_name = str(row.get("case_name") or "").strip()
            if not case_name:
                rejected += 1
                continue
            completed_year = str(row.get("completed_year") or infer_completed_year(case_name, str(row.get("source_title") or ""), str(row.get("published_at") or "")))
            inferred_completed = infer_is_completed(
                case_name,
                str(row.get("source_title") or ""),
                str(row.get("why") or ""),
                str(row.get("deal_status") or ""),
            )
            is_completed = parse_bool(row.get("is_completed"), default=inferred_completed)
            if not is_completed:
                rejected += 1
                continue
            row["completed_year"] = completed_year
            row["is_completed"] = str(is_completed).lower()
            rows.append(row)  # type: ignore[arg-type]
            accepted += 1
        LOGGER.info("Raw candidate LLM summary chunk done: chunk=%s accepted=%s rejected=%s accumulated=%s", chunk_index, accepted, rejected, len(rows))
        if len(rows) >= target_count:
            break
    return rows_to_briefs(rows)


def dedupe_briefs(briefs: list[CaseBrief]) -> list[CaseBrief]:
    seen_keys: set[str] = set()
    out: list[CaseBrief] = []
    for brief in briefs:
        if not brief.is_allowed_topic() or not has_explicit_parties(brief) or is_excluded_case(brief):
            continue
        keys = case_identity_keys(brief) or {brief.key()}
        if keys_overlap(keys, seen_keys):
            continue
        seen_keys.update(keys)
        out.append(brief)
    return out


def discover_backfill_cases(target_count: int, *, include_live: bool | None = None) -> list[CaseBrief]:
    if include_live is None:
        include_live = os.getenv("REPORT_BACKFILL_LIVE_SEARCH", "0") == "1"
    raw: list[RawItem] = []
    if include_live:
        end = datetime.now(BEIJING_TZ).replace(microsecond=0)
        start = end - timedelta(days=730)
        from mna_weekly_tracker.sources_rich import fetch_bing_news, fetch_google_news

        for query in CASE_DISCOVERY_QUERIES:
            try:
                raw.extend(fetch_google_news(query, start, end, source_name="Google News - backfill", source_url="https://news.google.com/", region_hint="中国"))
                raw.extend(fetch_bing_news(query, start, end, source_name="Bing News - backfill", source_url="https://www.bing.com/news/search", region_hint="中国"))
            except Exception as exc:  # noqa: BLE001
                LOGGER.warning("Backfill live discovery query failed and will be skipped: query=%s error=%s", query, exc)
    live_briefs = summarize_raw_items(raw, target_count=max(target_count, 20)) if raw else []
    pooled = live_briefs + extended_pool_briefs() + seed_briefs()
    deduped = dedupe_briefs(pooled)
    LOGGER.info("Backfill candidate pool: live=%s pool=%s seeds=%s deduped=%s target=%s", len(live_briefs), len(extended_pool_briefs()), len(seed_briefs()), len(deduped), target_count)
    return deduped


def choose_balanced(briefs: list[CaseBrief], *, count: int = 4, min_domestic: int = 2, report_root: Path, readiness_first: bool = False) -> list[CaseBrief]:
    deduped = dedupe_briefs(briefs)
    historical_keys = historical_case_keys(report_root)
    before_history_filter = len(deduped)
    history_kept: list[CaseBrief] = []
    history_rejected: list[CaseBrief] = []
    for brief in deduped:
        if any_key_in_history(case_identity_keys(brief), historical_keys):
            history_rejected.append(brief)
        else:
            history_kept.append(brief)
    deduped = history_kept
    LOGGER.info(
        "Historical duplicate filter: before=%s after=%s rejected=%s report_root=%s sample=%s",
        before_history_filter,
        len(deduped),
        len(history_rejected),
        report_root,
        " | ".join(brief.case_name for brief in history_rejected[:8]),
    )
    source_linked_completed_pool = [brief for brief in deduped if is_report_source_linked_completed_candidate(brief)]
    if os.getenv("REPORT_READY_ONLY", "1") == "1":
        before_ready_filter = len(deduped)
        ready_rejections = Counter(report_rejection_reason(brief) for brief in deduped if not is_report_ready_candidate(brief))
        deduped = [brief for brief in deduped if is_report_ready_candidate(brief)]
        LOGGER.info("Strict report-ready filter: before=%s after=%s rejected=%s", before_ready_filter, len(deduped), dict(ready_rejections))
    if os.getenv("REPORT_SOURCE_READY_ONLY", "1") == "1":
        before_source_filter = len(deduped)
        source_ready = [brief for brief in deduped if is_report_source_ready_candidate(brief)]
        source_rejections = Counter(report_rejection_reason(brief) for brief in deduped if not is_report_source_ready_candidate(brief))
        if (
            len(source_ready) < count
            and os.getenv("REPORT_ALLOW_WEAK_SOURCE_CANDIDATES", "0") == "1"
        ):
            weak_source = [
                brief for brief in source_linked_completed_pool
                if is_report_source_linked_completed_candidate(brief) and not is_report_source_ready_candidate(brief)
            ]
            deduped = dedupe_briefs(source_ready + weak_source)
            LOGGER.info(
                "Strict source-ready pool below requested selected count; adding weak source-linked completed backups: before=%s source_ready=%s weak_source=%s after=%s rejected=%s",
                before_source_filter,
                len(source_ready),
                len(weak_source),
                len(deduped),
                dict(source_rejections),
            )
        else:
            deduped = source_ready
            LOGGER.info("Strict report-source-ready filter: before=%s after=%s rejected=%s", before_source_filter, len(deduped), dict(source_rejections))
    counts = existing_counts(report_root)
    selected: list[CaseBrief] = []
    selected_keys: set[str] = set()
    selected_category_counts: Counter[str] = Counter()

    def score(b: CaseBrief) -> tuple[int, int, int, int, int, int, int, int, int, int, int, str]:
        preflight_penalty = 0 if is_report_ready_candidate(b) else 20
        selected_category_penalty = selected_category_counts[b.category] * 100
        existing_category_penalty = counts[b.category] + (2 if b.category == "SPAC" else 0)
        topic_rank = 0 if b.is_recent_completed() else 1
        domestic_rank = 0 if b.is_domestic else 1
        classic_penalty = 1 if b.is_classic else 0
        quality = report_candidate_priority(b)
        return (
            preflight_penalty,
            selected_category_penalty,
            existing_category_penalty,
            quality[0],
            quality[1],
            quality[2],
            quality[3],
            topic_rank,
            domestic_rank,
            classic_penalty,
            quality[4],
            quality[6],
        )

    if readiness_first or count <= 1:
        single_candidates = [b for b in deduped if b.is_domestic or min_domestic <= 0]
        if not single_candidates:
            single_candidates = deduped
        ordered = sorted(single_candidates, key=lambda brief: (0 if is_report_ready_candidate(brief) else 20, report_candidate_priority(brief)))[:count]
        LOGGER.info(
            "Selected %s single-report candidates by readiness, including %s domestic cases, requested count=%s min_domestic=%s",
            len(ordered),
            sum(1 for x in ordered if x.is_domestic),
            count,
            min_domestic,
        )
        for order, brief in enumerate(ordered, start=1):
            LOGGER.info(
                "Report candidate order %s: ready=%s case=%s category=%s completed=%s value=%s url=%s",
                order,
                is_report_ready_candidate(brief),
                brief.case_name,
                brief.category,
                brief.is_completed,
                brief.deal_value or "-",
                brief.source_url or "-",
            )
        return ordered

    # First pass: prefer one case per category, starting from historically underrepresented folders.
    for category in sorted(CATEGORIES, key=lambda c: counts[c] + (2 if c == "SPAC" else 0)):
        if len(selected) >= count:
            break
        category_candidates = [b for b in deduped if b.category == category and not keys_overlap(case_identity_keys(b) or {b.key()}, selected_keys)]
        if not category_candidates:
            continue
        brief = sorted(category_candidates, key=score)[0]
        selected.append(brief)
        selected_keys.update(case_identity_keys(brief) or {brief.key()})
        selected_category_counts[brief.category] += 1
        counts[brief.category] += 1

    # Second pass: fill any remaining slots, still penalizing categories already selected in this run.
    for brief in sorted([b for b in deduped if not keys_overlap(case_identity_keys(b) or {b.key()}, selected_keys)], key=score):
        if len(selected) >= count:
            break
        selected.append(brief)
        selected_keys.update(case_identity_keys(brief) or {brief.key()})
        selected_category_counts[brief.category] += 1
        counts[brief.category] += 1

    domestic_now = sum(1 for x in selected if x.is_domestic)
    domestic_backup_target = min(count, max(min_domestic, min_domestic + int(os.getenv("REPORT_DOMESTIC_BACKUP_CANDIDATES", "5")))) if min_domestic > 0 else 0
    if domestic_now < domestic_backup_target:
        for brief in sorted([b for b in deduped if b.is_domestic and not keys_overlap(case_identity_keys(b) or {b.key()}, selected_keys)], key=score):
            if domestic_now >= domestic_backup_target:
                break
            replace_indexes = sorted(
                [i for i, item in enumerate(selected) if not item.is_domestic],
                key=lambda i: (
                    selected_category_counts[selected[i].category],
                    not is_report_ready_candidate(selected[i]),
                    report_candidate_priority(selected[i]),
                ),
                reverse=True,
            )
            if not replace_indexes:
                break
            idx = replace_indexes[0]
            old = selected[idx]
            selected_category_counts[old.category] -= 1
            selected_keys.difference_update(case_identity_keys(old) or {old.key()})
            selected[idx] = brief
            selected_keys.update(case_identity_keys(brief) or {brief.key()})
            selected_category_counts[brief.category] += 1
            domestic_now = sum(1 for x in selected if x.is_domestic)

    LOGGER.info(
        "Selected %s report cases, including %s domestic cases, categories=%s, requested count=%s min_domestic=%s",
        len(selected[:count]),
        sum(1 for x in selected[:count] if x.is_domestic),
        dict(Counter(x.category for x in selected[:count])),
        count,
        min_domestic,
    )
    ordered = sorted(selected[:count], key=lambda brief: (0 if is_report_ready_candidate(brief) else 20, report_candidate_priority(brief)))
    for order, brief in enumerate(ordered, start=1):
        LOGGER.info(
            "Report candidate order %s: ready=%s case=%s category=%s completed=%s value=%s url=%s",
            order,
            is_report_ready_candidate(brief),
            brief.case_name,
            brief.category,
            brief.is_completed,
            brief.deal_value or "-",
            brief.source_url or "-",
        )
    return ordered


def save_manifest(path: Path, briefs: list[CaseBrief]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps([b.to_dict() for b in briefs], ensure_ascii=False, indent=2), encoding="utf-8")
