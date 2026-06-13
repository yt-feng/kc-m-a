"""Excel writer for weekly M&A cases."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

from .config import ALL_TRACKED_SOURCES, CATEGORIES, GLOBAL_QUERIES, HKEX_QUERIES, MIDDLE_EAST_QUERIES, OUTPUT_COLUMNS
from .sources import RawItem
from .sources_fixed import is_aggregator_url, is_likely_homepage_url, unwrap_news_url

HEADER_FILL = "0B7D73"
HEADER_FONT = "FFFFFF"
CATEGORY_FILL = "FFF2CC"
BORDER_COLOR = "B7B7B7"
HYPERLINK_COLOR = "0563C1"
URL_RE = re.compile(r"https?://[^\s；;，,。)）\]】>\"']+")

COLUMN_WIDTHS = {
    "A": 24, "B": 8, "C": 24, "D": 28, "E": 28, "F": 26, "G": 26, "H": 52,
    "I": 16, "J": 18, "K": 12, "L": 36, "M": 26, "N": 52, "O": 20, "P": 16,
}


def safe_cell(value: object) -> object:
    if isinstance(value, str) and value.startswith("="):
        return "'" + value
    return value


def first_url(value: object) -> str:
    """Return the first real http(s) URL contained in a cell value.

    Excel supports only one hyperlink target per cell. When a cell contains
    multiple URLs separated by Chinese semicolons, use the first valid URL as
    the click target while keeping the full display text in the cell.
    """
    if value is None:
        return ""
    match = URL_RE.search(str(value).strip())
    return clean_url(match.group(0).strip()) if match else ""


def clean_url(value: object) -> str:
    url = unwrap_news_url(str(value or "").strip())
    return "" if is_aggregator_url(url) or is_likely_homepage_url(url) else url


def make_url_cell_clickable(cell) -> None:
    url = first_url(cell.value)
    if not url:
        return
    cell.hyperlink = url
    cell.font = Font(color=HYPERLINK_COLOR, underline="single")


def apply_url_hyperlinks(ws, header_names: tuple[str, ...] = ("URL",)) -> None:
    """Convert URL text columns into real Excel hyperlinks."""
    header_to_col = {str(cell.value or "").strip(): cell.column for cell in ws[1]}
    for header in header_names:
        col_idx = header_to_col.get(header)
        if not col_idx:
            continue
        for row_idx in range(2, ws.max_row + 1):
            make_url_cell_clickable(ws.cell(row=row_idx, column=col_idx))


def style_sheet(ws) -> None:
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
    category_fill = PatternFill(fill_type="solid", fgColor=CATEGORY_FILL)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color=HEADER_FONT, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        row[0].fill = category_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter] = ColumnDimension(ws, index=letter, width=COLUMN_WIDTHS.get(letter, 18))
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 44 if row_idx > 1 else 26


def build_workbook(cases: list[dict[str, str]], raw_items: list[RawItem], errors: list[str], *, start_label: str, end_label: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "周度并购案例"
    ws.append(OUTPUT_COLUMNS)
    for idx, row in enumerate(cases, start=1):
        ws.append([idx if col == "序号" else safe_cell(clean_url(row.get(col, "")) if col == "URL" else row.get(col, "-")) for col in OUTPUT_COLUMNS])
    style_sheet(ws)
    apply_url_hyperlinks(ws)

    meta = wb.create_sheet("运行摘要")
    meta_rows = [
        ["字段", "内容"],
        ["统计区间", f"{start_label} 至 {end_label}（北京时间，最近 1 周）"],
        ["结构化案例数", len(cases)],
        ["原始候选数", len(raw_items)],
        ["生成时间", datetime.now().isoformat(timespec="seconds")],
        ["A列分类口径", "；".join(CATEGORIES)],
        ["提示", "如 DeepSeek 返回 '-'，表示候选来源未披露或无法可靠判断，建议人工复核。"],
    ]
    if errors:
        meta_rows.append(["采集警告", "\n".join(errors[:20])])
    for r in meta_rows:
        meta.append(r)
    for cell in meta[1]:
        cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
        cell.font = Font(color=HEADER_FONT, bold=True)
        cell.alignment = Alignment(horizontal="center")
    meta.column_dimensions["A"].width = 18
    meta.column_dimensions["B"].width = 120
    for row in meta.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sources = wb.create_sheet("跟踪信息源")
    sources.append(["来源/查询名称", "来源类型", "覆盖范围", "交易阶段", "建议频率", "优先级", "URL", "关键词/查询"])
    for source in ALL_TRACKED_SOURCES:
        sources.append([source.name, source.kind, source.coverage, source.stage, source.frequency, source.priority or "-", source.url, "；".join(source.keywords)])
    sources.append(["Google News - Global M&A", "google_news_rss", "全球并购新闻", "宣告/进展/交割", "每周", "-", "https://news.google.com/", "；".join(GLOBAL_QUERIES)])
    sources.append(["Google/Bing News - Middle East outbound M&A", "news_search", "中东主权基金、政府控股平台和产业资本收购/入股海外企业", "早期线索/宣告/完成", "每周", "P2", "https://news.google.com/；https://www.bing.com/news/search?format=rss", "；".join(MIDDLE_EAST_QUERIES)])
    sources.append(["Google/Bing News - HKEXnews", "news_search", "港股中国企业并购、收购、私有化、主要交易", "公告/股东批准/完成", "每周", "P1", "https://www.hkexnews.hk/", "；".join(HKEX_QUERIES)])
    for cell in sources[1]:
        cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
        cell.font = Font(color=HEADER_FONT, bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for idx, width in enumerate([36, 18, 38, 20, 16, 10, 58, 86], start=1):
        sources.column_dimensions[get_column_letter(idx)].width = width
    for row in sources.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    apply_url_hyperlinks(sources)

    raw = wb.create_sheet("原始候选")
    raw.append(["标题", "来源名称", "发布时间", "地区", "查询词", "URL", "摘要"])
    for item in raw_items:
        raw.append([item.title, item.source_name, item.published_at, item.region_hint, item.query, clean_url(item.url), item.summary])
    for cell in raw[1]:
        cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
        cell.font = Font(color=HEADER_FONT, bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for col, width in {"A": 60, "B": 30, "C": 22, "D": 14, "E": 28, "F": 70, "G": 80}.items():
        raw.column_dimensions[col].width = width
    raw.freeze_panes = "A2"
    raw.auto_filter.ref = raw.dimensions
    for row in raw.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    apply_url_hyperlinks(raw)
    return wb


def save_workbook(wb: Workbook, output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
