"""CLI entrypoint for weekly M&A case tracking."""

from __future__ import annotations

import argparse
import json
import logging
import os
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .config import OUTPUT_COLUMNS
from .deepseek import case_identities, case_identity, identities_overlap, structure_cases
from .excel import build_workbook, save_workbook
from .sources_rich import RawItem, fetch_all_candidates, week_window
from .url_validation import validate_workbook
from .weekly_windows import (
    DEFAULT_RECOVERY_START_DATE,
    WeeklyWindow,
    latest_scheduled_cutoff,
    missing_weekly_windows,
    window_from_dates,
)

LOGGER = logging.getLogger(__name__)


def configure_logging(verbose: bool = False) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(asctime)s %(levelname)s %(name)s - %(message)s")


def label(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d %H:%M")


def parse_raw_json(path: str | Path) -> list[RawItem]:
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    return [RawItem(**row) for row in data]


def load_previous_case_identities(output_dir: Path, *, current_output: Path | None = None) -> set[str]:
    identities: set[str] = set()
    for workbook_path in sorted(output_dir.glob("并购案例一览_*.xlsx")):
        if current_output and workbook_path.resolve() == current_output.resolve():
            continue
        try:
            wb = load_workbook(workbook_path, read_only=True, data_only=True)
            ws = wb["周度并购案例"] if "周度并购案例" in wb.sheetnames else wb.active
            rows = ws.iter_rows(values_only=True)
            headers = [str(x or "") for x in next(rows, [])]
            header_map = {name: idx for idx, name in enumerate(headers)}
            for values in rows:
                row = {
                    col: str(values[header_map[col]] or "")
                    for col in OUTPUT_COLUMNS
                    if col in header_map and header_map[col] < len(values)
                }
                identities.update(case_identities(row))
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("Failed to read previous workbook for duplicate filtering: %s error=%s", workbook_path, exc)
    return identities


def filter_previous_cases(cases: list[dict[str, str]], previous_identities: set[str]) -> list[dict[str, str]]:
    if not previous_identities:
        return cases
    filtered: list[dict[str, str]] = []
    skipped = 0
    for row in cases:
        identity = case_identity(row)
        identities = case_identities(row)
        if identities and identities_overlap(identities, previous_identities):
            skipped += 1
            LOGGER.info("Skipping previously exported case: %s -> %s", row.get("并购方", "-"), row.get("目标方", "-"))
            continue
        filtered.append(row)
    LOGGER.info("Filtered %s previously exported cases from weekly Excel output", skipped)
    return filtered


def resolve_target_windows(args: argparse.Namespace) -> list[WeeklyWindow]:
    if bool(args.start_date) != bool(args.end_date):
        raise ValueError("--start-date and --end-date must be provided together")
    if args.start_date and args.end_date:
        return [window_from_dates(args.start_date, args.end_date)]
    if args.recover_missing_weeks:
        if args.days != 7:
            raise ValueError("automatic weekly recovery requires --days 7; use explicit dates for another range")
        latest_end = latest_scheduled_cutoff()
        windows = missing_weekly_windows(
            Path(args.output_dir),
            latest_end,
            recovery_start_date=args.recovery_start_date,
        )
        LOGGER.info(
            "Weekly recovery plan cutoff=%s recovery_start=%s missing_count=%s windows=%s",
            label(latest_end),
            args.recovery_start_date,
            len(windows),
            [f"{window.start:%Y%m%d}_{window.end:%Y%m%d}" for window in windows],
        )
        return windows
    start, end = week_window(args.days, "Asia/Shanghai")
    return [WeeklyWindow(start=start, end=end)]


def write_generated_list(path: str, generated_paths: list[Path]) -> None:
    if not path:
        return
    list_path = Path(path)
    list_path.parent.mkdir(parents=True, exist_ok=True)
    text = "".join(f"{generated_path.as_posix()}\n" for generated_path in generated_paths)
    list_path.write_text(text, encoding="utf-8")


def generate_window(args: argparse.Namespace, window: WeeklyWindow) -> Path:
    start = window.start
    end = window.end
    start_label = label(start)
    end_label = label(end)
    output_dir = Path(args.output_dir)
    output_path = window.output_path(output_dir)
    LOGGER.info(
        "Generating weekly window start=%s end=%s output=%s",
        start_label,
        end_label,
        output_path,
    )
    if args.raw_json:
        LOGGER.info("Loading raw candidates from %s", args.raw_json)
        raw_items = parse_raw_json(args.raw_json)
        errors: list[str] = []
    else:
        LOGGER.info("Collecting candidates for %s to %s", start_label, end_label)
        raw_items, errors = fetch_all_candidates(start, end, max_items=args.max_raw_items)
    LOGGER.info("Collected %s raw candidates", len(raw_items))
    cases = structure_cases(raw_items, start_label=start_label, end_label=end_label, max_cases=args.max_cases)
    previous_identities = load_previous_case_identities(output_dir, current_output=output_path)
    cases = filter_previous_cases(cases, previous_identities)
    LOGGER.info("Structured %s cases", len(cases))
    wb = build_workbook(cases, raw_items, errors, start_label=start_label, end_label=end_label)
    save_workbook(wb, output_path)

    validation = validate_workbook(output_path)
    if validation["issue_count"] or validation["case_rows"] <= 0:
        issue_types = sorted({str(issue.get("type", "unknown")) for issue in validation["issues"]})
        raise RuntimeError(
            f"Generated workbook failed completeness validation: path={output_path} "
            f"case_rows={validation['case_rows']} issues={issue_types}"
        )
    LOGGER.info(
        "Wrote complete weekly workbook path=%s case_rows=%s url_count=%s",
        output_path,
        validation["case_rows"],
        validation["url_count"],
    )
    print(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Collect and structure weekly global M&A cases into Excel.")
    parser.add_argument(
        "--days",
        type=int,
        default=int(os.getenv("MNA_LOOKBACK_DAYS", "7")),
        help="Lookback window in days. Default: 7.",
    )
    parser.add_argument(
        "--output-dir",
        default=os.getenv("MNA_OUTPUT_DIR", "outputs"),
        help="Directory to write weekly Excel files.",
    )
    parser.add_argument(
        "--max-raw-items",
        type=int,
        default=int(os.getenv("MAX_RAW_ITEMS", "450")),
        help="Maximum raw candidates to keep. Default: 450.",
    )
    parser.add_argument(
        "--max-cases",
        type=int,
        default=int(os.getenv("MAX_STRUCTURED_CASES", "120")),
        help="Maximum structured cases in the Excel. Default: 120.",
    )
    parser.add_argument("--raw-json", default="", help="Optional local raw candidates JSON for debugging without network collection.")
    parser.add_argument("--start-date", default="", help="Explicit weekly start date in YYYY-MM-DD format; requires --end-date.")
    parser.add_argument("--end-date", default="", help="Explicit weekly end date in YYYY-MM-DD format; requires --start-date.")
    parser.add_argument(
        "--recover-missing-weeks",
        action="store_true",
        help="Generate every missing canonical weekly workbook through the latest Friday 05:00 cutoff.",
    )
    parser.add_argument(
        "--recovery-start-date",
        default=os.getenv("MNA_RECOVERY_START_DATE", DEFAULT_RECOVERY_START_DATE),
        help=f"Continuity baseline for automatic recovery. Default: {DEFAULT_RECOVERY_START_DATE}.",
    )
    parser.add_argument(
        "--write-generated-list",
        default="",
        help="Write newline-delimited paths for workbooks generated by this invocation.",
    )
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    configure_logging(args.verbose)
    write_generated_list(args.write_generated_list, [])
    try:
        windows = resolve_target_windows(args)
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc
    if args.raw_json and len(windows) != 1:
        raise SystemExit("--raw-json requires exactly one target window")

    generated_paths: list[Path] = []
    for index, window in enumerate(windows, start=1):
        LOGGER.info(
            "Weekly generation progress window=%s/%s start=%s end=%s",
            index,
            len(windows),
            window.start.date(),
            window.end.date(),
        )
        generated_paths.append(generate_window(args, window))
        write_generated_list(args.write_generated_list, generated_paths)

    if args.recover_missing_weeks:
        remaining = missing_weekly_windows(
            Path(args.output_dir),
            latest_scheduled_cutoff(),
            recovery_start_date=args.recovery_start_date,
        )
        if remaining:
            ranges = [f"{window.start:%Y%m%d}_{window.end:%Y%m%d}" for window in remaining]
            raise RuntimeError(f"Weekly coverage remains incomplete after recovery: {ranges}")
        LOGGER.info("Weekly coverage continuity check passed; no missing or invalid workbooks remain")

    write_generated_list(args.write_generated_list, generated_paths)
    LOGGER.info("Generated weekly workbook count=%s paths=%s", len(generated_paths), generated_paths)


if __name__ == "__main__":
    main()
